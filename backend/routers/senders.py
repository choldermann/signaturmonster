import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel
from typing import Optional
from database import get_db
from models import SenderProfile
from routers.auth import get_current_user_id

# CSV/XLSX column aliases: maps header variants → model field
_COL_MAP = {
    "email": "email",
    "e-mail": "email",
    "vorname": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "nachname": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "berufsbezeichnung": "job_title",
    "position": "job_title",
    "job_title": "job_title",
    "jobtitle": "job_title",
    "firma": "company",
    "company": "company",
    "telefon": "phone",
    "phone": "phone",
    "mobil": "mobile",
    "mobile": "mobile",
    "strasse": "street",
    "street": "street",
    "plz": "postal_code",
    "postal_code": "postal_code",
    "postalcode": "postal_code",
    "ort": "city",
    "city": "city",
    "land": "country",
    "country": "country",
    "foto_url": "photo_url",
    "photo_url": "photo_url",
    "photourl": "photo_url",
}

_EXPORT_COLUMNS = [
    ("email",       "email"),
    ("first_name",  "first_name"),
    ("last_name",   "last_name"),
    ("job_title",   "job_title"),
    ("company",     "company"),
    ("phone",       "phone"),
    ("mobile",      "mobile"),
    ("street",      "street"),
    ("postal_code", "postal_code"),
    ("city",        "city"),
    ("country",     "country"),
    ("photo_url",   "photo_url"),
]

def _parse_rows(rows: list[dict]) -> list[dict]:
    """Normalize column headers and return cleaned row dicts."""
    out = []
    for row in rows:
        mapped = {}
        for k, v in row.items():
            field = _COL_MAP.get(k.strip().lower().replace("-", "_"))
            if field:
                mapped[field] = str(v).strip() if v is not None else ""
        if mapped.get("email"):
            out.append(mapped)
    return out

def _rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")  # strip BOM
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]

def _rows_from_xlsx(content: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    return [dict(zip(headers, (str(c) if c is not None else "" for c in row))) for row in rows[1:]]

router = APIRouter()

class SenderIn(BaseModel):
    email: str
    first_name: str = ""
    last_name: str = ""
    job_title: str = ""
    photo_url: str = ""
    phone: str = ""
    mobile: str = ""
    street: str = ""
    postal_code: str = ""
    city: str = ""
    country: str = ""
    company: str = ""

class SenderOut(BaseModel):
    id: int
    email: str
    first_name: str = ""
    last_name: str = ""
    job_title: str = ""
    photo_url: str = ""
    phone: str = ""
    mobile: str = ""
    street: str = ""
    postal_code: str = ""
    city: str = ""
    country: str = ""
    company: str = ""
    claimed_by_user_id: Optional[int] = None
    class Config:
        from_attributes = True

class SelfServiceUpdate(BaseModel):
    first_name: str = ""
    last_name: str = ""
    job_title: str = ""
    photo_url: str = ""
    phone: str = ""
    mobile: str = ""
    street: str = ""
    postal_code: str = ""
    city: str = ""
    country: str = ""
    company: str = ""

# ── Self-Service (eigenes Profil) ─────────────────────────────────────────────

@router.get("/unclaimed", response_model=list[SenderOut])
async def list_unclaimed(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(SenderProfile)
        .where(SenderProfile.claimed_by_user_id == None)
        .order_by(SenderProfile.last_name, SenderProfile.first_name)
    )
    return result.scalars().all()

@router.get("/me", response_model=Optional[SenderOut])
async def get_my_profile(
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(SenderProfile).where(SenderProfile.claimed_by_user_id == user_id)
    )
    return result.scalar_one_or_none()

@router.post("/me/claim/{sender_id}", response_model=SenderOut)
async def claim_profile(
    sender_id: int,
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    existing = await db.execute(
        select(SenderProfile).where(SenderProfile.claimed_by_user_id == user_id)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Du hast bereits ein verknüpftes Profil")
    result = await db.execute(select(SenderProfile).where(SenderProfile.id == sender_id))
    sender = result.scalar_one_or_none()
    if not sender:
        raise HTTPException(404, "Profil nicht gefunden")
    if sender.claimed_by_user_id is not None:
        raise HTTPException(409, "Dieses Profil ist bereits verknüpft")
    sender.claimed_by_user_id = user_id
    await db.commit()
    await db.refresh(sender)
    return sender

@router.delete("/me/claim")
async def release_claim(
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(SenderProfile).where(SenderProfile.claimed_by_user_id == user_id)
    )
    sender = result.scalar_one_or_none()
    if not sender:
        raise HTTPException(404, "Kein verknüpftes Profil")
    sender.claimed_by_user_id = None
    await db.commit()
    return {"ok": True}

@router.put("/me", response_model=SenderOut)
async def update_my_profile(
    data: SelfServiceUpdate,
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(SenderProfile).where(SenderProfile.claimed_by_user_id == user_id)
    )
    sender = result.scalar_one_or_none()
    if not sender:
        raise HTTPException(404, "Kein verknüpftes Profil")
    for k, v in data.model_dump().items():
        setattr(sender, k, v)
    await db.commit()
    await db.refresh(sender)
    return sender

# ── Admin CRUD ────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[SenderOut])
async def list_senders(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SenderProfile).order_by(SenderProfile.last_name, SenderProfile.first_name))
    return result.scalars().all()

@router.post("/", response_model=SenderOut)
async def create_sender(data: SenderIn, db: AsyncSession = Depends(get_db)):
    sender = SenderProfile(**data.model_dump())
    db.add(sender)
    await db.commit()
    await db.refresh(sender)
    return sender

@router.put("/{sender_id}", response_model=SenderOut)
async def update_sender(sender_id: int, data: SenderIn, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SenderProfile).where(SenderProfile.id == sender_id))
    sender = result.scalar_one_or_none()
    if not sender:
        raise HTTPException(404, "Sender not found")
    for k, v in data.model_dump().items():
        setattr(sender, k, v)
    await db.commit()
    await db.refresh(sender)
    return sender

@router.delete("/{sender_id}")
async def delete_sender(sender_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SenderProfile).where(SenderProfile.id == sender_id))
    sender = result.scalar_one_or_none()
    if not sender:
        raise HTTPException(404, "Sender not found")
    await db.delete(sender)
    await db.commit()
    return {"ok": True}

@router.get("/by-email/{email}", response_model=Optional[SenderOut])
async def get_by_email(email: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SenderProfile).where(SenderProfile.email == email))
    return result.scalar_one_or_none()

# ── Import ────────────────────────────────────────────────────────────────────

@router.post("/import")
async def import_senders(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _user_id: int = Depends(get_current_user_id),
):
    content = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        rows = _rows_from_xlsx(content)
    else:
        rows = _rows_from_csv(content)

    parsed = _parse_rows(rows)
    created = updated = 0
    errors: list[str] = []

    for row in parsed:
        email = row.get("email", "").lower()
        if not email:
            continue
        try:
            result = await db.execute(select(SenderProfile).where(SenderProfile.email == email))
            existing = result.scalar_one_or_none()
            if existing:
                for k, v in row.items():
                    if k != "email" and v:
                        setattr(existing, k, v)
                updated += 1
            else:
                db.add(SenderProfile(**{k: v for k, v in row.items() if v}))
                created += 1
        except Exception as e:
            errors.append(f"{email}: {e}")

    await db.commit()
    return {"created": created, "updated": updated, "errors": errors, "total": len(parsed)}

# ── Export ────────────────────────────────────────────────────────────────────

@router.get("/export")
async def export_senders(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
    _user_id: int = Depends(get_current_user_id),
):
    result = await db.execute(select(SenderProfile).order_by(SenderProfile.last_name, SenderProfile.first_name))
    senders = result.scalars().all()
    headers_row = [col for col, _ in _EXPORT_COLUMNS]

    if format == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Absender-Profile"
        ws.append(headers_row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FCE499")
            cell.alignment = Alignment(horizontal="center")
        for s in senders:
            ws.append([getattr(s, field, "") or "" for _, field in _EXPORT_COLUMNS])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=absender-profile.xlsx"},
        )

    # CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers_row)
    for s in senders:
        writer.writerow([getattr(s, field, "") or "" for _, field in _EXPORT_COLUMNS])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=absender-profile.csv"},
    )
